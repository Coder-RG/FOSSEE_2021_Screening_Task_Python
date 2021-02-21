import sqlite3
import view_database as vdb
import openpyxl

dbname = "steel_sections.sqlite"

class NotValidDB(Exception):
    pass

class ItemNotStored(Exception):
    pass

class ItemAlreadyStored(Exception):
    pass

class NotValidTable(Exception):
    pass

def start_connection():
    """Establish connection with Database"""
    global dbname
    conn = sqlite3.connect(dbname)
    return conn

def connect(func):
    def inner(conn, *args, **kwargs):
        query = 'SELECT name FROM sqlite_master WHERE type="table";'
        try:
            conn.execute(query)
        except Exception as e:
            conn = start_connection()
        return func(conn, *args, **kwargs)
    return inner

def end_connection(conn):
    return

@connect
def insert_one(conn, table_name, values):
    """Inserts data into table mentioned in the `table_name`"""
    if table_name == 'Beams':
        #20
        query = 'INSERT INTO Beams VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    elif table_name == 'Angles':
        #24
        query = 'INSERT INTO Angles VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    else:
        #21
        query = 'INSERT INTO Channels VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'

    query2 = "SELECT designation FROM {} WHERE designation=?".format(table_name)
    present = conn.execute(query2, (values[0],)).fetchone()
    if present:
        raise ItemAlreadyStored()
    id = get_Id(conn, table_name)
    values.insert(0, int(id)+1)
    conn.execute(query, values)

@connect
def insert_upon_delete(conn, table_name, values):
    """Updates the value by first deleting the previous value and then inserting new."""
    if table_name == 'Beams':
        #20
        query = 'INSERT INTO Beams VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    elif table_name == 'Angles':
        #24
        query = 'INSERT INTO Angles VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    else:
        #21
        query = 'INSERT INTO Channels VALUES\
        (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);'
    conn.execute(query, values)

@connect
def view_all(conn, table_name):
    """View Table data of `table_name`"""
    if table_name is None:
        raise NotValidTable("Enter a valid table name")
    query = "SELECT * FROM {}".format(table_name)
    return conn.execute(query)

@connect
def view_one(conn, table_name, designation):
    if table_name is None:
        raise NotValidTable("Enter a valid table name")
    query = "SELECT * FROM {} WHERE designation=?".format(table_name)
    return conn.execute(query, (designation,))

@connect
def update_one(conn, table_name, values):
    query = "SELECT Id FROM {} WHERE designation=?".format(table_name)
    result = conn.execute(query, (values[0],)).fetchone()
    query2 = "DELETE FROM {} WHERE Id=?".format(table_name)
    conn.execute(query2, result)
    values.insert(0, result[0])
    insert_upon_delete(conn, table_name, values)

@connect
def delete_one(conn, table_name, designation):
    query = "DELETE FROM {} WHERE designation=?".format(table_name)
    conn.execute(query, (designation,))

@connect
def get_columns(conn, table_name):
    column_names = list()
    for row in conn.execute('PRAGMA table_info="{}"'.format(table_name)):
        column_names.append(row[1])
    return column_names

@connect
def get_designations(conn, table, *args, **kwargs):
    result = list()
    query = "SELECT Designation from {}".format(table)
    for i in conn.execute(query):
        result.append(i[0])
    return result

def get_Id(conn, table, *args, **kwargs):
    query = "SELECT MAX(Id) from {}".format(table)
    result = conn.execute(query).fetchone()
    return result[0]

if __name__ == '__main__':
    conn = start_connection()
    for row in conn.execute('SELECT * FROM Beams;'):
        print(row)
    # end_connection(conn)
    # wb = openpyxl.load_workbook(loc)
    # sheet = wb.active
    # max_col = sheet.max_column
    # values = [sheet.cell(row=2, column=i).value for i in range(1, max_col+1)]
    # insert_into_table(None, 'steel_sections.sqlite', 'Beams', values)
